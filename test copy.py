import pandas as pd
import openpyxl 
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the Excel file
file_path = 'sample_order_data.xlsx'
print("File path:", os.path.abspath(file_path))
wb = load_workbook(filename=file_path)

# Specify the worksheet name where the data resides
worksheet_names = wb.sheetnames

for worksheet_name in worksheet_names:
    # Read data from the worksheet
    df = pd.read_excel(file_path, sheet_name=worksheet_name, header=2) #added a header to start reading 2 rows down

    #print(f"Examining worksheet: {worksheet_name}")
    #print("Columns:", df.columns)

    # df edits  
    if worksheet_name == 'order_line_items':

        # Read required sheets
        products_df = pd.read_excel(file_path, sheet_name='products', header=2)
        order_line_items_df = pd.read_excel(file_path, sheet_name='order_line_items', header=2)
        orders_df = pd.read_excel(file_path, sheet_name='orders', header=2)
        customers_df = pd.read_excel(file_path, sheet_name='customers', header=2)
        salespersons_df = pd.read_excel(file_path, sheet_name='salespersons', header=2)

        print("Orders DataFrame:")
        print(orders_df.head())    
        print("Customers DataFrame:")
        print(customers_df.head()) 

        # Get the active worksheet
        ws = wb[worksheet_name]

        #print("Products DataFrame:")
        #print(products_df.head())

        #print("Order Line Items DataFrame:")
        #print(order_line_items_df.head())

        #print("Orders DataFrame:")
        #print(orders_df.head())

    # Merge necessary DataFrames
        merged_price_df = pd.merge(order_line_items_df, products_df[['product_id', 'product_price']], on='product_id', how='left')

        # Assign the "product_price" values to the "item_price" column
        merged_price_df['item_price'] = merged_price_df['product_price']

        # Calculate total units
        merged_price_df['total_units'] = merged_price_df['quantity_ordered'] - merged_price_df['quantity_canceled']


        # Calculate line totals
        merged_price_df['line_total'] = merged_price_df['product_price'] * merged_price_df['total_units']
        
        # Calculate total line total for each order
        order_total_df = merged_price_df.groupby('order_id')['line_total'].agg('sum').reset_index()
        order_total_df['total_units'] = order_total_df['line_total']
        
    # Reasign columns to DataFrames
        
        order_line_items_df['item_price'] = merged_price_df['item_price']
        order_line_items_df['line_total'] = merged_price_df['line_total']
        order_line_items_df['total_units'] = merged_price_df['total_units']

        orders_df['order_total'] = order_total_df['line_total']       

        #print("Merged Price DataFrame:")
        #print(merged_price_df.head())

        #print("Order Total DataFrame:")
        #print(order_total_df.head())

        #print("Order Line Items Updated DataFrame:")
        #print(order_line_items_df.head())

        #print("Orders Updated DataFrame:")
        #print(orders_df.head())

    #Write data to sheets
        ws = wb['order_line_items']

        order_line_items_item_price = order_line_items_df.columns.get_loc('item_price')
        for idx, value in enumerate(order_line_items_df['item_price'], start=4):
            ws.cell(row=idx, column=order_line_items_item_price + 1).value = value

        order_line_items_item_price = order_line_items_df.columns.get_loc('line_total')
        for idx, value in enumerate(order_line_items_df['line_total'], start=4):
            ws.cell(row=idx, column=order_line_items_item_price + 1).value = value

        order_line_items_item_price = order_line_items_df.columns.get_loc('total_units')
        for idx, value in enumerate(order_line_items_df['total_units'], start=4):
            ws.cell(row=idx, column=order_line_items_item_price + 1).value = value

        #print("Order Line Items Updated DataFrame:")
        #print(order_line_items_df.head())

        ws = wb['orders']
        orders_total_df = orders_df.columns.get_loc('order_total')
        for idx, value in enumerate(orders_df['order_total'], start=4):
            ws.cell(row=idx, column=orders_total_df + 1).value = value    

        # Merge necessary DataFrames
        orders_df = pd.merge(order_line_items_df, orders_df, on='order_id',)

        #print("Orders Updated DataFrame:")
        #print(orders_df.head())

        orders_df['order_date'] = pd.to_datetime(orders_df['order_date'])
        order_totals = orders_df.groupby('order_id')['order_total'].sum().reset_index()

        monthly_sales = orders_df.groupby(orders_df['order_date'].dt.to_period('M'))['line_total'].sum()

        monthly_sales = monthly_sales.reset_index()

        monthly_sales.rename(columns={'order_date': 'date', 'line_total': 'sales'}, inplace=True)

        #print("Monthly Sales DataFrame:")
        #print(monthly_sales.head())

        monthly_sales['date'] = monthly_sales['date'].dt.strftime('%Y-%m-%d')

        # Create or switch to Monthly_sales Sheet
        sheet_name = 'monthly_sales'
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(title=sheet_name)
            
        ws = wb['monthly_sales']
        for r_idx, row in enumerate(dataframe_to_rows(monthly_sales, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value,)  

        for col in ws.iter_cols():
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col[0].column_letter].width = adjusted_width   
        
        # Create the 'total_spent' column
        ws = wb['customers']
        
        if 'total_spent' not in customers_df.columns:
            customers_df['total_spent'] = 0
                    # Find the index of the last column
            last_column_index = ws.max_column
            # Insert a new column after the last column
            ws.insert_cols(idx=last_column_index + 1)
            # Set the header of the new column
            ws.cell(row=3, column=last_column_index + 1, value='total_spent')

        customer_total_df = pd.merge(orders_df, customers_df, on='customer_id', how='left')

        #print("Customer Total DataFrame:")
        #print(customer_total_df.head())     

        #print("Orders DataFrame:")
        #print(orders_df.head())    
        
        #print("Customers DataFrame:")
        #print(customers_df.head())    
        

        total_spent_df = orders_df.groupby('customer_id')['line_total'].sum().reset_index()

                # Rename columns for clarity
        total_spent_df.rename(columns={'line_total': 'total_spent'}, inplace=True)

        print("Orders DataFrame:")
        print(orders_df.head())  

        print("Total Spent DataFrame:")
        print(total_spent_df.head())    

        #print("Customers DataFrame:")
        #print(customers_df.head())        

        #print("Customer Total DataFrame:")
        #print(customer_total_df.head())  

        total_spent_column_index = customers_df.columns.get_loc('total_spent')
        for idx, value in enumerate(total_spent_df['total_spent'], start=4):
            ws.cell(row=idx, column=total_spent_column_index + 1).value = value

# Save the changes to the Excel file
wb.save(file_path)
