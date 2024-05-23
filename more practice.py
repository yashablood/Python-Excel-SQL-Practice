import pandas as pd
import openpyxl 
import os
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the Excel file
file_path = 'sample_order_data.xlsx'
print("File path:", os.path.abspath(file_path))
wb = load_workbook(filename=file_path)

# Specify the worksheet name where the data resides
worksheet_names = wb.sheetnames

for worksheet_name in worksheet_names:
    # Read data from the worksheet
    df = pd.read_excel(file_path, sheet_name=worksheet_name, header=2) #added a header to start reading 2 rows down

    #print(f"Examining worksheet: {worksheet_name}")
    #print("Columns:", df.columns)

    # df edits  
    if worksheet_name == 'order_line_items':
        # Read required sheets
        products_df = pd.read_excel(file_path, sheet_name='products', header=2)
        order_line_items_df = pd.read_excel(file_path, sheet_name='order_line_items', header=2)
        orders_df = pd.read_excel(file_path, sheet_name='orders', header=2)

        # Get the active worksheet
        ws = wb[worksheet_name]

        print("Products DataFrame:")
        print(products_df.head())

        print("Order Line Items DataFrame:")
        print(order_line_items_df.head())

        print("Orders DataFrame:")
        print(orders_df.head())

    # Merge necessary DataFrames
        merged_price_df = pd.merge(order_line_items_df, products_df[['product_id', 'product_price']], on='product_id', how='left')

        # Assign the "product_price" values to the "item_price" column
        merged_price_df['item_price'] = merged_price_df['product_price']

        # Calculate total units
        merged_price_df['total_units'] = merged_price_df['quantity_ordered'] - merged_price_df['quantity_canceled']


        # Calculate line totals
        merged_price_df['line_total'] = merged_price_df['product_price'] * merged_price_df['total_units']
        
        # Calculate total line total for each order
        order_total_df = merged_price_df.groupby('order_id')['line_total'].agg('sum').reset_index()
        order_total_df['total_units'] = order_total_df['line_total']
        
    # Reasign columns to DataFrames
        
        order_line_items_df['item_price'] = merged_price_df['item_price']
        order_line_items_df['line_total'] = merged_price_df['line_total']
        order_line_items_df['total_units'] = merged_price_df['total_units']

        orders_df['order_total'] = order_total_df['line_total']       

        print("Merged Price DataFrame:")
        print(merged_price_df.head())

        print("Order Total DataFrame:")
        print(order_total_df.head())

        print("Order Line Items Updated DataFrame:")
        print(order_line_items_df.head())

        print("Orders Updated DataFrame:")
        print(orders_df.head())

    #Write data to sheets
        ws = wb['order_line_items']
        order_line_items_item_price = order_line_items_df.columns.get_loc('item_price')
        for idx, value in enumerate(order_line_items_df['item_price'], start=4):
            ws.cell(row=idx, column=order_line_items_item_price + 1).value = value
        order_line_items_item_price = order_line_items_df.columns.get_loc('line_total')
        for idx, value in enumerate(order_line_items_df['line_total'], start=4):
            ws.cell(row=idx, column=order_line_items_item_price + 1).value = value
        order_line_items_item_price = order_line_items_df.columns.get_loc('total_units')
        for idx, value in enumerate(order_line_items_df['total_units'], start=4):
            ws.cell(row=idx, column=order_line_items_item_price + 1).value = value

        print("Order Line Items Updated DataFrame:")
        print(order_line_items_df.head())

        ws = wb['orders']
        orders_total_df = orders_df.columns.get_loc('order_total')
        for idx, value in enumerate(orders_df['order_total'], start=4):
            ws.cell(row=idx, column=orders_total_df + 1).value = value    

# Save the changes to the Excel file
wb.save(file_path)
